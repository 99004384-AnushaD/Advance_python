'''
   Autor: Anusha D
   PS number: 99004384
   project description: Reading the Excel Sheet and searching the data
'''

"""
class
"""
import os
import openpyxl
def getrequiredata(dic1, req_ps):
    """
    :param dic1:
    :param req_ps:
    :return:
    """
    requireddict = {}
    for key, value in dic1.items():
        for dic in value:
            for keysdata in dic.keys():
                if req_ps == dic[keysdata]:
                    requireddict[key] = dic
    return requireddict

def list2dict(fulllist):
    """"
    :param fulllist: It is a list contains N number of Lists
    :return: Returns a list contains N number of dictionaries having
    keys : first list of fulllist
    values : second to last of fulllist as values
    """
    dic_list = []
    for f_list in range(1, len(fulllist)):
        res = dict(zip(fulllist[0], fulllist[f_list]))
        dic_list.append(res)
    return dic_list

def load_excel(path):
    """
    :param path:
    :return:
    """
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    return wb_obj

def getexceldata(sheetobj, maxrow, maxcol):
    """
    :param sheetobj:
    :param maxrow:
    :param maxcol:
    :return:
    """
    full_list = []
    for rows in range(1, maxrow + 1):
        st_dat = []
        for col in range(1, maxcol + 1):
            cell_obj = sheetobj.cell(row=rows, column=col).value
            st_dat.append(cell_obj)
        full_list.append(st_dat)
    return full_list


full_list = []


class Student:
    """
    Student class used to get the data from the excel
    """

    def __init__(self):
        """
        constructor
        """
        self.path = "python-data.xlsx"

    def main(self):
        """
        :return:
        """
        while True:
            req_ps = int(input("Enter the PS Number(between 99004351 and 99004365): "))
            if req_ps < 99004351 or req_ps > 99004365:
                print("The entered PS_Number: " + str(req_ps) + " is not available. ")
                continue

            break

        dic1 = {}
        try:
            xlobj = load_excel(self.path)
        except FileNotFoundError:
            print('File not found!!!!', self.path)

        xlobj = load_excel(self.path)
        sns = xlobj.sheetnames
        while True:
            req_sheetname = str(input("Enter the sheet name in the given list:" + str(sns) + "\n"))
            if req_sheetname not in sns:
                print("The entered sheet name: " + str(req_sheetname) + " is not available. ")
                continue
            break
        filepath = "Output.xlsx"
        work_book = openpyxl.Workbook()
        # wb.create_sheet(index=1, title="demo sheet2")
        sheet = work_book.active
        for sheetnumber in sns:
            sheet1 = xlobj[sheetnumber]
            maxcolumnss = self.getmaxcolumn(xlobj, sheetnumber)
            maxrowss = self.getmaxrow(xlobj, sheetnumber)
            res = getexceldata(sheet1, maxrowss, maxcolumnss)
            out = list2dict(res)
            dic1[sheetnumber] = out
            data = getrequiredata(dic1, req_ps)
            if len(data.items()) == 0:
                print("PS No: " + str(req_ps) + " Not Found in Sheet name " + sheetnumber)
                break
            rowvalue = 0
            for sheetkey, sheetdatavalue in data.items():
                for key, value in sheetdatavalue.items():

                    if req_sheetname == sheetkey:
                        rowvalue += 1
                        sheet.cell(row=rowvalue, column=1).value = sheetkey
                        sheet.cell(row=rowvalue, column=2).value = key
                        sheet.cell(row=rowvalue, column=3).value = value
        print(data[req_sheetname])
        work_book.save(filepath)
        print("**********************************************************************")
        print(filepath + ' is generated. Available in location: "' + os.path.join(os.getcwd(), filepath) + '"')
        print("**********************************************************************")

    def getmaxcolumn(self, xlobj, requiredsheetname):
        """
        :param xlobj:
        :param requiredsheetname:
        :return:
        """
        columnsnum = 0
        sheetnames = xlobj.sheetnames
        for sheetname in sheetnames:
            if sheetname == requiredsheetname:
                sheet1 = xlobj[sheetname]
                columnsnum = sheet1.max_column
        return columnsnum

    def getmaxrow(self, xlobj, requiredsheetname):
        """
        :param xlobj:
        :param requiredsheetname:
        :return:
        """
        rowssnum = 0
        sheetnames = xlobj.sheetnames
        for sheetname in sheetnames:
            if sheetname == requiredsheetname:
                sheet1 = xlobj[sheetname]
                rowssnum = sheet1.max_row
        return rowssnum

if __name__ == "__main__":
    student = Student()  # s is an reference variable which is refering to student object
    student.main()
