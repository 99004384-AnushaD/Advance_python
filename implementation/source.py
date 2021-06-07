from openpyxl import workbook, load_workbook

wb = load_workbook('python_project_DATA.xlsx')    

class file:
  

    def __init__(self, filename):
        self.fileName = filename


    def read(self, filename):
        df = pd.read_excel(filename, sheet_name=[0, 1, 2, 3, 4])
        return df

class PopulateExcel(file):

    def save(self, filename, output):

        # load_workbook( ) function is used
        # when you have to access an MS Excel file in openpyxl module.
       

        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
        writer.book = book

        # ExcelWriter for some reason uses
        # writer.sheets to access the sheet.
       

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        cols = []
        for j in range(len(output.columns)):
            cols.append(output.columns[j])

        output.to_excel(writer, "newsheet", columns=cols, index=False)

        writer.save()
        
ws = wb.active           

def my_ps_number(ip_ws):
    ps_nums=[]
    for row in ws.iter_rows(min_row=1, max_col=1, max_row=16, values_only=True):
        ps_nums.append(list(row))
    return ps_nums

def show_ps_number(ps):
    for item in ps:
        print(item)

def main():
    print("\nEnter the PS number from the below list:\n")
    ps = my_ps_number(ws)
    show_ps_number(ps)
    user_choice = int(input("\nEnter the Ps number:"))

main()
